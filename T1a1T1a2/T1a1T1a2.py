from __future__ import annotations
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple
import numpy as np
try:
    from openpyxl import Workbook
    HAVE_OPENPYXL = True
except Exception:
    Workbook = None
    HAVE_OPENPYXL = False
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
try:
    from mpi4py import MPI
    COMM = MPI.COMM_WORLD
    RANK = COMM.Get_rank()
    SIZE = COMM.Get_size()
except Exception:
    COMM = None
    RANK = 0
    SIZE = 1
FONT_TITLE = 17
FONT_LABEL = 19
FONT_TICK_AXIS = 19
FONT_TICK_CBAR = 12
VIEW_ELEV = 28
VIEW_AZIM = -120
THREE_D_CMAP = 'jet'
SINGULAR_EPS = 1e-10

def barrier() -> None:
    if COMM is not None:
        COMM.Barrier()
N_PARAM = 100
ALPHA_MIN = 0.0
ALPHA_MAX = 2.0
BETA_MIN = 0.0
BETA_MAX = 1.0
TOPO_NK = 500
TOPO_GAMMA = 0.0
SHG_NKX = 1400
SHG_NKY = 1400
OMEGA = 0.5
ETA_PV1 = 0.01
ETA_PV2 = 0.01
USE_WILSON_LOOP = True
OUTDIR = 'phase_shg_outputs_mpi_v4_fullSHG'
PHASE_TOL = 1e-08
SAVE_NPY = True
SAVE_ALL_SHG_HEATMAPS = True

@dataclass
class Config:
    n_param: int = N_PARAM
    alpha_min: float = ALPHA_MIN
    alpha_max: float = ALPHA_MAX
    beta_min: float = BETA_MIN
    beta_max: float = BETA_MAX
    topo_nk: int = TOPO_NK
    topo_gamma: float = TOPO_GAMMA
    shg_nkx: int = SHG_NKX
    shg_nky: int = SHG_NKY
    omega: float = OMEGA
    eta_pv1: float = ETA_PV1
    eta_pv2: float = ETA_PV2
    use_wilson_loop: bool = USE_WILSON_LOOP
    outdir: str = OUTDIR
    phase_tol: float = PHASE_TOL
    save_npy: bool = SAVE_NPY
    save_all_shg_heatmaps: bool = SAVE_ALL_SHG_HEATMAPS

def save_heatmap_shg(x_vec: np.ndarray, y_vec: np.ndarray, data_beta_alpha: np.ndarray, title: str, fpath: Path) -> None:
    fig = plt.figure(figsize=(6.5, 5.2))
    ax = fig.add_subplot(111)
    im = ax.pcolormesh(x_vec, y_vec, data_beta_alpha, shading='auto', cmap='jet')
    ax.set_xlabel('$\\alpha$', fontsize=FONT_LABEL)
    ax.set_ylabel('$\\beta$', fontsize=FONT_LABEL)
    ax.tick_params(axis='both', labelsize=FONT_TICK_AXIS)
    cbar = fig.colorbar(im, ax=ax)
    cbar.ax.tick_params(labelsize=FONT_TICK_CBAR)
    fig.tight_layout()
    fig.savefig(fpath, dpi=300)
    plt.close(fig)

def save_3d_surface_shg(x_vec: np.ndarray, y_vec: np.ndarray, data_beta_alpha: np.ndarray, title: str, fpath: Path, elev: float=VIEW_ELEV, azim: float=VIEW_AZIM) -> None:
    X, Y = np.meshgrid(x_vec, y_vec)
    Z = np.asarray(data_beta_alpha, dtype=float)
    fig = plt.figure(figsize=(7.2, 6.0))
    ax = fig.add_subplot(111, projection='3d')
    surf = ax.plot_surface(X, Y, Z, cmap=THREE_D_CMAP, linewidth=0.35, edgecolor='k', antialiased=True)
    ax.set_xlabel('$\\alpha$', fontsize=FONT_LABEL, labelpad=8)
    ax.set_ylabel('$\\beta$', fontsize=FONT_LABEL, labelpad=8)
    ax.tick_params(axis='both', labelsize=FONT_TICK_AXIS, pad=4)
    ax.tick_params(axis='z', labelsize=FONT_TICK_AXIS, pad=10)
    ax.zaxis.set_tick_params(pad=10)
    ax.view_init(elev=elev, azim=azim)
    cbar = fig.colorbar(surf, ax=ax, shrink=0.72, pad=0.08)
    cbar.ax.tick_params(labelsize=FONT_TICK_CBAR)
    fig.subplots_adjust(left=0.02, right=0.88, bottom=0.06, top=0.94)
    fig.savefig(fpath, dpi=300, bbox_inches='tight', pad_inches=0.15)
    plt.close(fig)

def grid_to_table(alpha_vec: np.ndarray, beta_vec: np.ndarray, grid_alpha_beta: np.ndarray) -> dict:
    return {'columns': [float(np.round(x, 12)) for x in alpha_vec], 'index': [float(np.round(y, 12)) for y in beta_vec], 'data': np.asarray(grid_alpha_beta.T, dtype=np.float64)}

def write_records_csv(rows: List[Dict[str, float]], filepath: Path) -> None:
    filepath.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        filepath.write_text('', encoding='utf-8-sig')
        return
    headers = list(rows[0].keys())
    with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

def write_grid_csv(table: dict, filepath: Path) -> None:
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['beta/alpha'] + table['columns'])
        for beta, row in zip(table['index'], table['data']):
            writer.writerow([beta] + list(row))

def add_records_sheet(wb, sheet_name: str, rows: List[Dict[str, float]]) -> None:
    ws = wb.create_sheet(title=sheet_name)
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, '') for h in headers])

def add_grid_sheet(wb, sheet_name: str, table: dict) -> None:
    ws = wb.create_sheet(title=sheet_name)
    ws.append(['beta/alpha'] + table['columns'])
    for beta, row in zip(table['index'], table['data']):
        ws.append([beta] + [float(x) if np.isfinite(x) else None for x in row])

def write_phase_summary_workbook(summary_rows: List[Dict[str, float]], filepath: Path) -> None:
    if not HAVE_OPENPYXL:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = 'phase_summary'
    headers = ['phase', 'point_count', 'T1a1_sum', 'T1a1_mean', 'T1a2_sum', 'T1a2_mean', 'Total_sum', 'Total_mean']
    ws.append(headers)
    for row in summary_rows:
        ws.append([row.get(h, '') for h in headers])
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)

class TopoPhaseCalculator:

    def __init__(self, cfg: Config):
        self.cfg = cfg
        self.alpha_vec = np.linspace(cfg.alpha_min, cfg.alpha_max, cfg.n_param)
        self.beta_vec = np.linspace(cfg.beta_min, cfg.beta_max, cfg.n_param)
        self.kx_vec = np.linspace(0.0, np.pi, cfg.topo_nk, endpoint=True)
        self.ky_vec = np.linspace(0.0, 2.0 * np.pi, cfg.topo_nk, endpoint=False)
        self.KX, self.KY = np.meshgrid(self.kx_vec, self.ky_vec, indexing='ij')

    def build_hamiltonian_grid(self, alpha: float, beta: float) -> np.ndarray:
        gamma = self.cfg.topo_gamma
        d1 = np.sin(self.KX) * np.cos(self.KY) + alpha * np.sin(self.KX) + beta * np.sin(2.0 * self.KX) + gamma * beta * np.sin(self.KX)
        d2 = np.sin(self.KX) * np.sin(self.KY)
        d3 = np.cos(self.KX)
        H = np.zeros((self.cfg.topo_nk, self.cfg.topo_nk, 2, 2), dtype=np.complex128)
        H[..., 0, 0] = d3
        H[..., 1, 1] = -d3
        H[..., 0, 1] = d1 - 1j * d2
        H[..., 1, 0] = d1 + 1j * d2
        return H

    def build_dHdkx_line(self, kx_val: float, alpha: float, beta: float) -> np.ndarray:
        gamma = self.cfg.topo_gamma
        dd1 = np.cos(kx_val) * np.cos(self.ky_vec) + alpha * np.cos(kx_val) + 2.0 * beta * np.cos(2.0 * kx_val) + gamma * beta * np.cos(kx_val)
        dd2 = np.cos(kx_val) * np.sin(self.ky_vec)
        dd3 = -np.sin(kx_val)
        dH = np.zeros((self.cfg.topo_nk, 2, 2), dtype=np.complex128)
        dH[:, 0, 0] = dd3
        dH[:, 1, 1] = -dd3
        dH[:, 0, 1] = dd1 - 1j * dd2
        dH[:, 1, 0] = dd1 + 1j * dd2
        return dH

    def calculate_eigensystem(self, alpha: float, beta: float) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
        H = self.build_hamiltonian_grid(alpha, beta)
        evals, evecs = np.linalg.eigh(H)
        u_v = evecs[..., 0]
        u_c = evecs[..., 1]
        return (evals, u_v, u_c)

    @staticmethod
    def calculate_rtpv_fukui(u_v_grid: np.ndarray) -> float:
        u1 = u_v_grid[:-1, :, :]
        u2 = u_v_grid[1:, :, :]
        u3 = np.roll(u_v_grid[1:, :, :], shift=-1, axis=1)
        u4 = np.roll(u_v_grid[:-1, :, :], shift=-1, axis=1)
        link_x1 = np.sum(np.conj(u1) * u2, axis=-1)
        link_y2 = np.sum(np.conj(u2) * u3, axis=-1)
        link_x2rev = np.sum(np.conj(u4) * u3, axis=-1)
        link_y1 = np.sum(np.conj(u1) * u4, axis=-1)
        plaquette = link_x1 * link_y2 * np.conj(link_x2rev) * np.conj(link_y1)
        total_flux = np.sum(np.angle(plaquette))
        return -total_flux / (2.0 * np.pi)

    @staticmethod
    def unit_phase_array(z: np.ndarray, eps: float=SINGULAR_EPS) -> np.ndarray:
        out = np.empty_like(z, dtype=np.complex128)
        mag = np.abs(z)
        mask = mag >= eps
        out[mask] = z[mask] / mag[mask]
        out[~mask] = np.nan + 1j * np.nan
        return out

    def calculate_S_line_wilson(self, kx_idx: int, evals: np.ndarray, u_v_grid: np.ndarray, u_c_grid: np.ndarray, alpha: float, beta: float, eps: float=SINGULAR_EPS) -> float:
        kx_val = self.kx_vec[kx_idx]
        uc = u_c_grid[kx_idx, :, :]
        uv = u_v_grid[kx_idx, :, :]
        uc_next = np.roll(uc, shift=-1, axis=0)
        uv_next = np.roll(uv, shift=-1, axis=0)
        link_c = np.sum(np.conj(uc) * uc_next, axis=-1)
        link_v = np.sum(np.conj(uv_next) * uv, axis=-1)
        dH = self.build_dHdkx_line(kx_val, alpha, beta)
        gap = evals[kx_idx, :, 1] - evals[kx_idx, :, 0]
        Ax_cv = np.einsum('ji,jik,jk->j', np.conj(uc), dH, uv) / (1j * gap)
        Ax_cv_next = np.roll(Ax_cv, shift=-1, axis=0)
        Ax_vc = np.conj(Ax_cv)
        if np.any(np.abs(gap) < eps) or np.any(np.abs(Ax_cv) < eps) or np.any(np.abs(Ax_cv_next) < eps) or np.any(np.abs(link_c) < eps) or np.any(np.abs(link_v) < eps):
            return np.nan
        W_phase = self.unit_phase_array(link_c, eps) * self.unit_phase_array(Ax_cv_next, eps) * self.unit_phase_array(link_v, eps) * self.unit_phase_array(Ax_vc, eps)
        if np.any(~np.isfinite(W_phase.real)) or np.any(~np.isfinite(W_phase.imag)):
            return np.nan
        total_angle = np.sum(np.angle(W_phase))
        return -total_angle / (2.0 * np.pi)

    @staticmethod
    def round_if_finite(x: float) -> float:
        return np.nan if not np.isfinite(x) else float(np.rint(x))

    def calculate_invariants_for_point(self, alpha: float, beta: float) -> np.ndarray:
        evals, u_v_grid, u_c_grid = self.calculate_eigensystem(alpha, beta)
        RTP_v = self.calculate_rtpv_fukui(u_v_grid)
        S_0 = self.calculate_S_line_wilson(0, evals, u_v_grid, u_c_grid, alpha, beta)
        S_pi = self.calculate_S_line_wilson(self.cfg.topo_nk - 1, evals, u_v_grid, u_c_grid, alpha, beta)
        if not np.isfinite(S_0) or not np.isfinite(S_pi):
            return np.array([np.nan, self.round_if_finite(RTP_v), np.nan, np.nan], dtype=np.float64)
        Vort_x = S_0 - S_pi - 2.0 * RTP_v
        return np.array([self.round_if_finite(Vort_x), self.round_if_finite(RTP_v), self.round_if_finite(S_0), self.round_if_finite(S_pi)], dtype=np.float64)

    def gather_results(self, all_results: List[List[Tuple[int, int, np.ndarray]]]) -> Dict[str, np.ndarray]:
        if RANK != 0:
            return {}
        na = len(self.alpha_vec)
        nb = len(self.beta_vec)
        invariants_grid = np.full((na, nb, 4), np.nan, dtype=np.float64)
        for sublist in all_results:
            for ia, ib, vals in sublist:
                invariants_grid[ia, ib, :] = vals
        vortx_grid = invariants_grid[:, :, 0]
        rtpv_grid = invariants_grid[:, :, 1]
        S_0_grid = invariants_grid[:, :, 2]
        S_pi_grid = invariants_grid[:, :, 3]
        minus_vortx_grid = -vortx_grid
        relative_shift_grid = S_0_grid - S_pi_grid
        average_shift_grid = 0.5 * (S_0_grid + S_pi_grid)
        return {'invariants_grid': invariants_grid, 'vortx_grid': vortx_grid, 'rtpv_grid': rtpv_grid, 'S_0_grid': S_0_grid, 'S_pi_grid': S_pi_grid, 'minus_vortx_grid': minus_vortx_grid, 'relative_shift_grid': relative_shift_grid, 'average_shift_grid': average_shift_grid}

class FullSHGCalculator:

    def __init__(self, cfg: Config):
        self.cfg = cfg
        self.Na = cfg.n_param
        self.Nb = cfg.n_param
        self.alpha_vec = np.linspace(cfg.alpha_min, cfg.alpha_max, self.Na)
        self.beta_vec = np.linspace(cfg.beta_min, cfg.beta_max, self.Nb)
        self.kx = np.linspace(-np.pi, np.pi, cfg.shg_nkx, endpoint=False)
        self.ky = np.linspace(0.0, 2.0 * np.pi, cfg.shg_nky, endpoint=False)
        self.d_kx = 2.0 * np.pi / cfg.shg_nkx
        self.d_ky = 2.0 * np.pi / cfg.shg_nky
        self.KX, self.KY = np.meshgrid(self.kx, self.ky, indexing='ij')

    @staticmethod
    def H_from_d(d: np.ndarray) -> np.ndarray:
        d1, d2, d3 = (d[..., 0], d[..., 1], d[..., 2])
        H = np.empty(d.shape[:-1] + (2, 2), dtype=complex)
        H[..., 0, 0] = d3
        H[..., 0, 1] = d1 - 1j * d2
        H[..., 1, 0] = d1 + 1j * d2
        H[..., 1, 1] = -d3
        return H

    @staticmethod
    def d_components(kx: np.ndarray, ky: np.ndarray, alpha: float, beta: float) -> np.ndarray:
        s, c = (np.sin, np.cos)
        d1 = s(kx) * c(ky) + alpha * s(kx) + beta * s(2 * kx)
        d2 = s(kx) * s(ky)
        d3 = c(kx)
        return np.stack([d1, d2, d3], axis=-1)

    @staticmethod
    def d_derivs(kx: np.ndarray, ky: np.ndarray, alpha: float, beta: float) -> Tuple[np.ndarray, np.ndarray]:
        s, c = (np.sin, np.cos)
        ax1 = c(kx) * c(ky) + alpha * c(kx) + 2 * beta * c(2 * kx)
        ax2 = c(kx) * s(ky)
        ax3 = -s(kx)
        a_x = np.stack([ax1, ax2, ax3], axis=-1)
        ay1 = -s(kx) * s(ky)
        ay2 = s(kx) * c(ky)
        ay3 = 0.0
        a_y = np.stack([ay1, ay2, np.full_like(ay1, ay3)], axis=-1)
        return (a_x, a_y)

    @staticmethod
    def V_from_a(a: np.ndarray) -> np.ndarray:
        a1, a2, a3 = (a[..., 0], a[..., 1], a[..., 2])
        V = np.empty(a.shape[:-1] + (2, 2), dtype=complex)
        V[..., 0, 0] = a3
        V[..., 0, 1] = a1 - 1j * a2
        V[..., 1, 0] = a1 + 1j * a2
        V[..., 1, 1] = -a3
        return V

    @staticmethod
    def kernel_plus(x: np.ndarray, eta: float) -> np.ndarray:
        return 1.0 / (x + 1j * eta / 2.0)

    @staticmethod
    def kernel_minus(x: np.ndarray, eta: float) -> np.ndarray:
        return 1.0 / (x - 1j * eta / 2.0)

    @staticmethod
    def eigh_2x2_herm_batch(H: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
        return np.linalg.eigh(H)

    @staticmethod
    def mat_elem(U_left: np.ndarray, Op: np.ndarray, U_right: np.ndarray) -> np.ndarray:
        left = np.conj(U_left)[..., None, :]
        right = U_right[..., :, None]
        return (left @ Op @ right)[..., 0, 0]

    @staticmethod
    def band_vectors(U: np.ndarray, band_idx: int) -> np.ndarray:
        return U[..., :, band_idx]

    @staticmethod
    def overlap(psi: np.ndarray, phi: np.ndarray) -> np.ndarray:
        return np.sum(np.conj(psi) * phi, axis=-1)

    @staticmethod
    def roll_plus(arr: np.ndarray, axis: int) -> np.ndarray:
        return np.roll(arr, -1, axis=axis)

    @staticmethod
    def roll_minus(arr: np.ndarray, axis: int) -> np.ndarray:
        return np.roll(arr, +1, axis=axis)

    def d_arg_central(self, z: np.ndarray, axis: int, dk: float) -> np.ndarray:
        z_plus = self.roll_plus(z, axis)
        z_minus = self.roll_minus(z, axis)
        up = z_plus / np.abs(z_plus)
        um = z_minus / np.abs(z_minus)
        return np.imag(np.log(up * np.conj(um) + 0j)) / (2.0 * dk)

    def d_logabs_central(self, z: np.ndarray, axis: int, dk: float) -> np.ndarray:
        z_plus = self.roll_plus(z, axis)
        z_minus = self.roll_minus(z, axis)
        return (np.log(np.abs(z_plus)) - np.log(np.abs(z_minus))) / (2.0 * dk)

    def berry_connection_forward(self, u: np.ndarray, axis: int, dk: float) -> np.ndarray:
        u_plus = self.roll_plus(u, axis)
        link = self.overlap(u, u_plus)
        return np.imag(np.log(link + 0j)) / dk

    def berry_connection_central(self, u: np.ndarray, axis: int, dk: float) -> np.ndarray:
        A_f = self.berry_connection_forward(u, axis, dk)
        A_b = -self.berry_connection_forward(self.roll_minus(u, axis), axis, dk)
        return 0.5 * (A_f + A_b)

    @staticmethod
    def r_from_v(v_cv: np.ndarray, omega_cv: np.ndarray) -> np.ndarray:
        return v_cv / (1j * omega_cv)

    def R_from_definition(self, vb_cv: np.ndarray, uc: np.ndarray, uv: np.ndarray, dk_a: float, axis_a: int) -> np.ndarray:
        dphi = self.d_arg_central(vb_cv, axis=axis_a, dk=dk_a)
        A_c = self.berry_connection_central(uc, axis=axis_a, dk=dk_a)
        A_v = self.berry_connection_central(uv, axis=axis_a, dk=dk_a)
        return -dphi + (A_c - A_v)

    def R_from_wilson(self, uc: np.ndarray, uv: np.ndarray, rb_cv: np.ndarray, dk_a: float, axis_a: int) -> np.ndarray:
        uc_p = self.roll_plus(uc, axis_a)
        uv_p = self.roll_plus(uv, axis_a)
        rb_p = self.roll_plus(rb_cv, axis_a)
        Wp = self.overlap(uc, uc_p) * rb_p * self.overlap(uv_p, uv) * np.conj(rb_cv)
        uc_m = self.roll_minus(uc, axis_a)
        uv_m = self.roll_minus(uv, axis_a)
        rb_m = self.roll_minus(rb_cv, axis_a)
        Wm = self.overlap(uc, uc_m) * rb_m * self.overlap(uv_m, uv) * np.conj(rb_cv)
        return -np.imag(np.log(Wp / (Wm + 0j) + 0j)) / (2.0 * dk_a)

    def subterms_on_kgrid_numeric(self, alpha: float, beta: float) -> Tuple[np.ndarray, ...]:
        kx = self.KX
        ky = self.KY
        d = self.d_components(kx, ky, alpha, beta)
        a_x, a_y = self.d_derivs(kx, ky, alpha, beta)
        H = self.H_from_d(d)
        Vx = self.V_from_a(a_x)
        Vy = self.V_from_a(a_y)
        E, U = self.eigh_2x2_herm_batch(H)
        uv = self.band_vectors(U, 0)
        uc = self.band_vectors(U, 1)
        Ev = E[..., 0]
        Ec = E[..., 1]
        omega_cv = Ec - Ev
        v_x_12 = self.mat_elem(uv, Vx, uc)
        v_y_12 = self.mat_elem(uv, Vy, uc)
        v_x_21 = np.conj(v_x_12)
        v_y_21 = np.conj(v_y_12)
        v_x_22 = self.mat_elem(uc, Vx, uc).real
        v_x_11 = self.mat_elem(uv, Vx, uv).real
        ratio = (v_x_22 - v_x_11) / omega_cv
        dkx = 2.0 * np.pi / self.cfg.shg_nkx
        dky = 2.0 * np.pi / self.cfg.shg_nky
        dln_abs_vx = self.d_logabs_central(v_x_12, axis=0, dk=dkx)
        if self.cfg.use_wilson_loop:
            r_x_12 = self.r_from_v(v_x_12, omega_cv)
            R_xx_12 = self.R_from_wilson(uv, uc, r_x_12, dk_a=dkx, axis_a=0)
            R_yx12 = self.R_from_wilson(uv, uc, r_x_12, dk_a=dky, axis_a=1)
        else:
            R_xx_12 = self.R_from_definition(v_x_12, uv, uc, dk_a=dkx, axis_a=0)
            R_yx12 = self.R_from_definition(v_x_12, uv, uc, dk_a=dky, axis_a=1)
        R_yx21 = -R_yx12
        x1_1 = (self.cfg.omega - omega_cv) / 2.0
        x1_2 = (-self.cfg.omega - omega_cv) / 2.0
        x2_1 = (2.0 * self.cfg.omega - omega_cv) / 2.0
        x2_2 = (-2.0 * self.cfg.omega - omega_cv) / 2.0
        K1_1 = self.kernel_plus(x1_1, self.cfg.eta_pv1)
        K1_2 = self.kernel_minus(x1_2, self.cfg.eta_pv1)
        K2_1 = self.kernel_plus(x2_1, self.cfg.eta_pv2)
        K2_2 = self.kernel_minus(x2_2, self.cfg.eta_pv2)
        axpn2 = np.abs(v_x_12) ** 2
        v_y12_x21 = v_y_12 * v_x_21
        v_y21_x12 = v_y_21 * v_x_12
        factor_T1a = -1j * axpn2 * R_yx21
        T1a_1_k = factor_T1a * K1_1
        T1a_2_k = factor_T1a * -1.0 * K1_2
        T1b_1_k = -1.0 * v_y12_x21 * ratio * K1_1
        T1b_2_k = -1.0 * v_y21_x12 * ratio * K1_2
        T2a_1_k = -1j * v_y12_x21 * R_xx_12 * K2_1
        T2a_2_k = +1j * v_y21_x12 * R_xx_12 * K2_2
        T2b_1_k = -1.0 * v_y12_x21 * dln_abs_vx * K2_1
        T2b_2_k = -1.0 * v_y21_x12 * dln_abs_vx * K2_2
        T2c_1_k = -1.0 * v_y12_x21 * ratio * K2_1
        T2c_2_k = -1.0 * v_y21_x12 * ratio * K2_2
        return (T1a_1_k, T1a_2_k, T1b_1_k, T1b_2_k, T2a_1_k, T2a_2_k, T2b_1_k, T2b_2_k, T2c_1_k, T2c_2_k)

    def integrate_point(self, ia: int, ib: int, alpha: float, beta: float) -> Tuple[int, int, Tuple[complex, ...]]:
        terms = self.subterms_on_kgrid_numeric(alpha, beta)
        dA = self.d_kx * self.d_ky
        pref1 = 1j / (4.0 * 4.0 * np.pi ** 2 * self.cfg.omega ** 3)
        pref2 = 1j / (8.0 * 4.0 * np.pi ** 2 * self.cfg.omega ** 3)
        I1a_1 = pref1 * (terms[0].sum() * dA)
        I1a_2 = pref1 * (terms[1].sum() * dA)
        I1b_1 = pref1 * (terms[2].sum() * dA)
        I1b_2 = pref1 * (terms[3].sum() * dA)
        I2a_1 = pref2 * (terms[4].sum() * dA)
        I2a_2 = pref2 * (terms[5].sum() * dA)
        I2b_1 = pref2 * (terms[6].sum() * dA)
        I2b_2 = pref2 * (terms[7].sum() * dA)
        I2c_1 = pref2 * (terms[8].sum() * dA)
        I2c_2 = pref2 * (terms[9].sum() * dA)
        return (ia, ib, (I1a_1, I1a_2, I1b_1, I1b_2, I2a_1, I2a_2, I2b_1, I2b_2, I2c_1, I2c_2))

    def gather_results(self, all_results: List[List[Tuple[int, int, Tuple[complex, ...]]]]) -> Dict[str, np.ndarray]:
        if RANK != 0:
            return {}
        keys_sub = ['T1a_1', 'T1a_2', 'T1b_1', 'T1b_2', 'T2a_1', 'T2a_2', 'T2b_1', 'T2b_2', 'T2c_1', 'T2c_2']
        keys_main = ['T1a', 'T1b', 'T2a', 'T2b', 'T2c']
        maps: Dict[str, np.ndarray] = {k: np.zeros((self.Nb, self.Na), dtype=np.complex128) for k in keys_sub + keys_main + ['Total']}
        for sublist in all_results:
            for ia, ib, vals in sublist:
                i1a1, i1a2, i1b1, i1b2, i2a1, i2a2, i2b1, i2b2, i2c1, i2c2 = vals
                maps['T1a_1'][ib, ia] = i1a1
                maps['T1a_2'][ib, ia] = i1a2
                maps['T1b_1'][ib, ia] = i1b1
                maps['T1b_2'][ib, ia] = i1b2
                maps['T2a_1'][ib, ia] = i2a1
                maps['T2a_2'][ib, ia] = i2a2
                maps['T2b_1'][ib, ia] = i2b1
                maps['T2b_2'][ib, ia] = i2b2
                maps['T2c_1'][ib, ia] = i2c1
                maps['T2c_2'][ib, ia] = i2c2
        maps['T1a'] = maps['T1a_1'] + maps['T1a_2']
        maps['T1b'] = maps['T1b_1'] + maps['T1b_2']
        maps['T2a'] = maps['T2a_1'] + maps['T2a_2']
        maps['T2b'] = maps['T2b_1'] + maps['T2b_2']
        maps['T2c'] = maps['T2c_1'] + maps['T2c_2']
        maps['Total'] = maps['T1a'] + maps['T1b'] + maps['T2a'] + maps['T2b'] + maps['T2c']
        return maps

def format_latex_subscript(key: str) -> str:
    prefix = 'yxx'
    if key == 'Total':
        return f'_{{{prefix}}}'
    if '_' in key:
        a, b = key.split('_')
        return f'_{{{prefix}, {a}_{b}}}'
    return f'_{{{prefix}, {key}}}'

def make_phase_masks(avg_shift_grid_alpha_beta: np.ndarray, tol: float) -> Dict[str, np.ndarray]:
    finite = np.isfinite(avg_shift_grid_alpha_beta)
    mask_zero = finite & np.isclose(avg_shift_grid_alpha_beta, 0.0, atol=tol)
    mask_mhalf = finite & np.isclose(avg_shift_grid_alpha_beta, -0.5, atol=tol)
    return {'$\\langle S \\rangle = 0$': mask_zero.T, '$\\langle S \\rangle = -1/2$': mask_mhalf.T}

def summarize_by_phase(phase_masks: Dict[str, np.ndarray], T1a1_abs_real: np.ndarray, T1a2_abs_real: np.ndarray) -> List[Dict[str, float]]:
    total_abs_real = T1a1_abs_real + T1a2_abs_real
    rows: List[Dict[str, float]] = []
    for phase_label, mask in phase_masks.items():
        count = int(np.count_nonzero(mask))
        if count == 0:
            rows.append({'phase': phase_label, 'point_count': 0, 'T1a1_sum': np.nan, 'T1a1_mean': np.nan, 'T1a2_sum': np.nan, 'T1a2_mean': np.nan, 'Total_sum': np.nan, 'Total_mean': np.nan})
            continue
        t1a1_sum = float(np.nansum(T1a1_abs_real[mask]))
        t1a2_sum = float(np.nansum(T1a2_abs_real[mask]))
        total_sum = float(np.nansum(total_abs_real[mask]))
        rows.append({'phase': phase_label, 'point_count': count, 'T1a1_sum': t1a1_sum, 'T1a1_mean': t1a1_sum / count, 'T1a2_sum': t1a2_sum, 'T1a2_mean': t1a2_sum / count, 'Total_sum': total_sum, 'Total_mean': total_sum / count})
    return rows

def distribution_stats_by_phase(phase_masks: Dict[str, np.ndarray], value_grid: np.ndarray, observable_name: str) -> List[Dict[str, float]]:
    rows: List[Dict[str, float]] = []
    for phase_label, mask in phase_masks.items():
        vals = value_grid[mask]
        vals = vals[np.isfinite(vals)]
        if vals.size == 0:
            rows.append({'observable': observable_name, 'phase': phase_label, 'count': 0, 'min': np.nan, 'q1': np.nan, 'median': np.nan, 'mean': np.nan, 'q3': np.nan, 'max': np.nan, 'std': np.nan})
        else:
            rows.append({'observable': observable_name, 'phase': phase_label, 'count': int(vals.size), 'min': float(np.min(vals)), 'q1': float(np.percentile(vals, 25)), 'median': float(np.percentile(vals, 50)), 'mean': float(np.mean(vals)), 'q3': float(np.percentile(vals, 75)), 'max': float(np.max(vals)), 'std': float(np.std(vals))})
    return rows

def _collect_phase_series(phase_masks: Dict[str, np.ndarray], value_grid: np.ndarray) -> Tuple[List[str], List[np.ndarray]]:
    labels = list(phase_masks.keys())
    series: List[np.ndarray] = []
    for label in labels:
        vals = value_grid[phase_masks[label]]
        vals = vals[np.isfinite(vals)]
        vals = vals[vals >= 0.0]
        if vals.size == 0:
            vals = np.array([np.nan])
        series.append(vals)
    return (labels, series)

def _robust_positive_bandwidth(vals: np.ndarray) -> float:
    vals = vals[np.isfinite(vals)]
    vals = vals[vals >= 0.0]
    n = vals.size
    if n < 2:
        return 1.0
    std = float(np.std(vals, ddof=1))
    iqr = float(np.percentile(vals, 75) - np.percentile(vals, 25))
    scale = min(std, iqr / 1.349) if iqr > 0 else std
    if not np.isfinite(scale) or scale <= 0:
        scale = max(float(np.max(vals) - np.min(vals)), 1.0)
    h = 0.9 * scale * n ** (-1.0 / 5.0)
    if not np.isfinite(h) or h <= 0:
        h = max(float(np.max(vals) - np.min(vals)) / 100.0, 1e-12)
    return float(h)

def _normalize_density_on_grid(density: np.ndarray, xs: np.ndarray) -> np.ndarray:
    density = np.asarray(density, dtype=float)
    density = np.where(np.isfinite(density), density, 0.0)
    density = np.maximum(density, 0.0)
    if xs.size < 2:
        return density
    area = float(np.trapezoid(density, xs))
    if np.isfinite(area) and area > 0:
        density = density / area
    return density

def _build_nonnegative_density_axis(series: List[np.ndarray]) -> np.ndarray:
    finite_series = []
    for vals in series:
        vv = vals[np.isfinite(vals)]
        vv = vv[vv >= 0.0]
        if vv.size > 0:
            finite_series.append(vv)
    if not finite_series:
        return np.array([])
    all_vals = np.concatenate(finite_series)
    xmax = float(np.max(all_vals))
    if xmax <= 0 or not np.isfinite(xmax):
        xmax = 1.0
    pad = 0.03 * xmax
    return np.linspace(0.0, xmax + pad, 700)

def _kde_density_nonnegative(vals: np.ndarray, xs: np.ndarray, chunk_size: int=256) -> np.ndarray:
    vals = vals[np.isfinite(vals)]
    vals = vals[vals >= 0.0]
    n = vals.size
    if n < 2:
        return np.zeros_like(xs, dtype=float)
    h = _robust_positive_bandwidth(vals)
    norm = 1.0 / (h * np.sqrt(2.0 * np.pi))
    dens = np.empty_like(xs, dtype=float)
    for i in range(0, xs.size, chunk_size):
        x_chunk = xs[i:i + chunk_size]
        z1 = (x_chunk[:, None] - vals[None, :]) / h
        z2 = (x_chunk[:, None] + vals[None, :]) / h
        dens[i:i + chunk_size] = np.mean(np.exp(-0.5 * z1 * z1) + np.exp(-0.5 * z2 * z2), axis=1) * norm
    return dens

def save_phase_hist_kde(phase_masks: Dict[str, np.ndarray], value_grid: np.ndarray, xlabel: str, filepath: Path, bins: int=40) -> None:
    labels, series = _collect_phase_series(phase_masks, value_grid)
    xs = _build_nonnegative_density_axis(series)
    if xs.size == 0:
        return
    fig = plt.figure(figsize=(8.2, 6.0))
    ax = fig.add_subplot(111)
    bin_edges = np.linspace(float(xs[0]), float(xs[-1]), bins + 1)
    for label, vals in zip(labels, series):
        vals = vals[np.isfinite(vals)]
        vals = vals[vals >= 0.0]
        if vals.size == 0:
            continue
        ax.hist(vals, bins=bin_edges, density=True, alpha=0.28, label=f'{label} histogram, n={vals.size}')
        density = _normalize_density_on_grid(_kde_density_nonnegative(vals, xs), xs)
        line, = ax.plot(xs, density, linewidth=2.0, label=f'{label} reflection KDE')
        color = line.get_color()
    ax.set_xlim(0.0, float(xs[-1]))
    ax.set_xlabel(xlabel, fontsize=FONT_LABEL)
    ax.set_ylabel('Probability density', fontsize=FONT_LABEL)
    ax.tick_params(axis='both', labelsize=FONT_TICK_AXIS)
    fig.tight_layout()
    fig.savefig(filepath, dpi=300)
    plt.close(fig)

def _get_ordered_phase_labels_for_two_row_plot(phase_masks: Dict[str, np.ndarray]) -> List[str]:
    labels = list(phase_masks.keys())
    label_mhalf = None
    label_zero = None
    for label in labels:
        if '-0.5' in label or '-1/2' in label:
            label_mhalf = label
        elif '= 0' in label or '=0' in label:
            label_zero = label
    ordered: List[str] = []
    if label_mhalf is not None:
        ordered.append(label_mhalf)
    if label_zero is not None:
        ordered.append(label_zero)
    for label in labels:
        if label not in ordered:
            ordered.append(label)
    return ordered

def _clean_phase_label_for_rows(label: str) -> str:
    if '-0.5' in label or '-1/2' in label:
        return '$\\langle S\\rangle=-1/2$'
    if '= 0' in label or '=0' in label:
        return '$\\langle S\\rangle=0$'
    return label

def _get_nonnegative_values_for_mask(value_grid: np.ndarray, mask: np.ndarray) -> np.ndarray:
    vals = value_grid[mask]
    vals = vals[np.isfinite(vals)]
    vals = vals[vals >= 0.0]
    return vals

def _phase_style_for_two_row_plot(label: str) -> Tuple[str, str, str, str]:
    if '-0.5' in label or '-1/2' in label:
        return ('#A1D99B', '#D62728', '#D62728', '$\\langle S\\rangle=-1/2$')
    if '= 0' in label or '=0' in label:
        return ('#9ECAE1', '#FF7F0E', '#FF7F0E', '$\\langle S\\rangle=0$')
    return ('#9ECAE1', '#FF7F0E', '#FF7F0E', _clean_phase_label_for_rows(label))

def save_phase_hist_kde_two_rows_single_observable(phase_masks: Dict[str, np.ndarray], value_grid: np.ndarray, xlabel: str, filepath: Path, bins: int=40) -> None:
    phase_labels = _get_ordered_phase_labels_for_two_row_plot(phase_masks)
    if len(phase_labels) < 2:
        return
    phase_labels = phase_labels[:2]
    series_for_axis: List[np.ndarray] = []
    for phase_label in phase_labels:
        vals = _get_nonnegative_values_for_mask(value_grid, phase_masks[phase_label])
        if vals.size > 0:
            series_for_axis.append(vals)
    if len(series_for_axis) == 0:
        return
    xs = _build_nonnegative_density_axis(series_for_axis)
    if xs.size == 0:
        return
    bin_edges = np.linspace(float(xs[0]), float(xs[-1]), bins + 1)
    fig, axes = plt.subplots(2, 1, figsize=(7.6, 8.6), sharex=True, sharey=False)
    for row, phase_label in enumerate(phase_labels):
        ax = axes[row]
        vals = _get_nonnegative_values_for_mask(value_grid, phase_masks[phase_label])
        hist_color, kde_color, mean_color, clean_label = _phase_style_for_two_row_plot(phase_label)
        if vals.size == 0:
            ax.text(0.5, 0.5, 'No data', ha='center', va='center', transform=ax.transAxes, fontsize=FONT_TICK_AXIS)
            continue
        ax.hist(vals, bins=bin_edges, density=True, alpha=0.3, color=hist_color, edgecolor='none')
        density = _normalize_density_on_grid(_kde_density_nonnegative(vals, xs), xs)
        ax.plot(xs, density, linewidth=2.4, color=kde_color)
        ax.set_xlim(0.0, float(xs[-1]))
        ax.tick_params(axis='both', labelsize=FONT_TICK_AXIS)
        ax.text(0.03, 0.94, clean_label, transform=ax.transAxes, ha='left', va='top', fontsize=FONT_LABEL)
        if row == 1:
            ax.set_xlabel(xlabel, fontsize=FONT_LABEL)
        else:
            ax.tick_params(labelbottom=False)
    fig.text(0.04, 0.5, 'Probability density', va='center', rotation='vertical', fontsize=FONT_LABEL)
    fig.subplots_adjust(left=0.18, right=0.98, bottom=0.1, top=0.98, hspace=0.08)
    fig.savefig(filepath, dpi=300)
    plt.close(fig)

def save_phase_kde_only(phase_masks: Dict[str, np.ndarray], value_grid: np.ndarray, xlabel: str, filepath: Path) -> None:
    labels, series = _collect_phase_series(phase_masks, value_grid)
    xs = _build_nonnegative_density_axis(series)
    if xs.size == 0:
        return
    fig = plt.figure(figsize=(8.2, 6.0))
    ax = fig.add_subplot(111)
    for label, vals in zip(labels, series):
        vals = vals[np.isfinite(vals)]
        vals = vals[vals >= 0.0]
        if vals.size == 0:
            continue
        density = _normalize_density_on_grid(_kde_density_nonnegative(vals, xs), xs)
        line, = ax.plot(xs, density, linewidth=2.2, label=f'{label} reflection KDE, n={vals.size}')
        color = line.get_color()
        ax.fill_between(xs, density, 0.0, alpha=0.18, color=color)
    ax.set_xlim(0.0, float(xs[-1]))
    ax.set_xlabel(xlabel, fontsize=FONT_LABEL)
    ax.set_ylabel('Probability density', fontsize=FONT_LABEL)
    ax.tick_params(axis='both', labelsize=FONT_TICK_AXIS)
    fig.tight_layout()
    fig.savefig(filepath, dpi=300)
    plt.close(fig)

def save_phase_mean_barplot(summary_rows: List[Dict[str, float]], mean_col: str, ylabel: str, filepath: Path, title: str) -> None:
    labels = [row['phase'] for row in summary_rows]
    values = [row.get(mean_col, np.nan) for row in summary_rows]
    fig = plt.figure(figsize=(7.2, 5.6))
    ax = fig.add_subplot(111)
    bars = ax.bar(labels, values)
    ax.set_ylabel(ylabel, fontsize=FONT_LABEL)
    ax.tick_params(axis='both', labelsize=FONT_TICK_AXIS)
    fig.tight_layout()
    fig.savefig(filepath, dpi=300)
    plt.close(fig)

def save_phase_masked_heatmap(x_vec: np.ndarray, y_vec: np.ndarray, value_grid_beta_alpha: np.ndarray, phase_mask_beta_alpha: np.ndarray, title: str, fpath: Path) -> None:
    masked = np.where(phase_mask_beta_alpha, value_grid_beta_alpha, np.nan)
    save_heatmap_shg(x_vec, y_vec, masked, title, fpath)

def build_phase_point_rows(phase_masks: Dict[str, np.ndarray], alpha_vec: np.ndarray, beta_vec: np.ndarray, average_shift_grid_alpha_beta: np.ndarray, T1a1_abs_real_beta_alpha: np.ndarray, T1a2_abs_real_beta_alpha: np.ndarray) -> List[Dict[str, float]]:
    rows: List[Dict[str, float]] = []
    total_abs = T1a1_abs_real_beta_alpha + T1a2_abs_real_beta_alpha
    for phase_label, mask in phase_masks.items():
        ib_ia_pairs = np.argwhere(mask)
        for ib, ia in ib_ia_pairs:
            rows.append({'phase': phase_label, 'ia': int(ia), 'ib': int(ib), 'alpha': float(alpha_vec[ia]), 'beta': float(beta_vec[ib]), 'average_shift': float(average_shift_grid_alpha_beta[ia, ib]), 'T1a1_abs_real': float(T1a1_abs_real_beta_alpha[ib, ia]), 'T1a2_abs_real': float(T1a2_abs_real_beta_alpha[ib, ia]), 'T1a_total_abs_real': float(total_abs[ib, ia])})
    return rows

def write_phase_point_workbook(rows: List[Dict[str, float]], filepath: Path) -> None:
    if not HAVE_OPENPYXL:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = 'phase_points'
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, '') for h in headers])
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)

def save_outputs(cfg: Config, topo_calc: TopoPhaseCalculator, shg_calc: FullSHGCalculator, topo_results: Dict[str, np.ndarray], shg_maps: Dict[str, np.ndarray], summary_rows: List[Dict[str, float]], distribution_rows: List[Dict[str, float]], phase_masks: Dict[str, np.ndarray]) -> None:
    outdir = Path(cfg.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    shg_dir = outdir / 'shg_maps'
    analysis_dir = outdir / 'phase_analysis_plots'
    tables_dir = outdir / 'tables'
    shg_dir.mkdir(parents=True, exist_ok=True)
    analysis_dir.mkdir(parents=True, exist_ok=True)
    tables_dir.mkdir(parents=True, exist_ok=True)
    alpha_vec = topo_calc.alpha_vec
    beta_vec = topo_calc.beta_vec
    average_shift_grid = topo_results['average_shift_grid']
    relative_shift_grid = topo_results['relative_shift_grid']
    minus_vortx_grid = topo_results['minus_vortx_grid']
    rtpv_grid = topo_results['rtpv_grid']
    S_0_grid = topo_results['S_0_grid']
    S_pi_grid = topo_results['S_pi_grid']
    vortx_grid = topo_results['vortx_grid']
    invariants_grid = topo_results['invariants_grid']
    real_dir = shg_dir / 'real'
    imag_dir = shg_dir / 'imag'
    mod_dir = shg_dir / 'mod'
    real3d_dir = shg_dir / 'real_3d'
    for d in [real_dir, imag_dir, mod_dir, real3d_dir]:
        d.mkdir(parents=True, exist_ok=True)
    keys_sub = ['T1a_1', 'T1a_2', 'T1b_1', 'T1b_2', 'T2a_1', 'T2a_2', 'T2b_1', 'T2b_2', 'T2c_1', 'T2c_2']
    keys_main = ['T1a', 'T1b', 'T2a', 'T2b', 'T2c']
    all_keys = keys_sub + keys_main + ['Total']
    if cfg.save_all_shg_heatmaps:
        for k in all_keys:
            label = format_latex_subscript(k)
            val = shg_maps[k]
            val_real_abs = np.abs(np.real(val))
            val_imag_abs = np.abs(np.imag(val))
            val_mod = np.abs(val)
            save_heatmap_shg(alpha_vec, beta_vec, val_real_abs, f'$|\\mathrm{{Re}}[\\chi^{{(2)}}{label}]|$', real_dir / f'{k}_real.png')
            save_heatmap_shg(alpha_vec, beta_vec, val_imag_abs, f'$|\\mathrm{{Im}}[\\chi^{{(2)}}{label}]|$', imag_dir / f'{k}_imag.png')
            save_heatmap_shg(alpha_vec, beta_vec, val_mod, f'$|\\chi^{{(2)}}{label}|$', mod_dir / f'{k}_mod.png')
    for k in ['T1a_1', 'T1a_2', 'T1a']:
        label = format_latex_subscript(k)
        val_real_abs = np.abs(np.real(shg_maps[k]))
        save_3d_surface_shg(alpha_vec, beta_vec, val_real_abs, f'$|\\mathrm{{Re}}[\\chi^{{(2)}}{label}]|$', real3d_dir / f'{k}_real_3d.png')
    T1a1_abs_real = np.abs(np.real(shg_maps['T1a_1']))
    T1a2_abs_real = np.abs(np.real(shg_maps['T1a_2']))
    total_abs_real = T1a1_abs_real + T1a2_abs_real
    save_phase_hist_kde(phase_masks, T1a1_abs_real, '$|\\mathrm{Re}[T1a_{1}]|$', analysis_dir / 'hist_kde_T1a1.png')
    save_phase_hist_kde(phase_masks, T1a2_abs_real, '$|\\mathrm{Re}[T1a_{2}]|$', analysis_dir / 'hist_kde_T1a2.png')
    save_phase_hist_kde(phase_masks, total_abs_real, '$|\\mathrm{Re}[T1a_{1}]|+|\\mathrm{Re}[T1a_{2}]|$', analysis_dir / 'hist_kde_T1a_total.png')
    save_phase_hist_kde_two_rows_single_observable(phase_masks, T1a1_abs_real, '$|\\mathrm{Re}[T1a_{1}]|$', analysis_dir / 'hist_kde_T1a1_by_phase_rows.png')
    save_phase_hist_kde_two_rows_single_observable(phase_masks, T1a2_abs_real, '$|\\mathrm{Re}[T1a_{2}]|$', analysis_dir / 'hist_kde_T1a2_by_phase_rows.png')
    save_phase_kde_only(phase_masks, T1a1_abs_real, '$|\\mathrm{Re}[T1a_{1}]|$', analysis_dir / 'kde_T1a1.png')
    save_phase_kde_only(phase_masks, T1a2_abs_real, '$|\\mathrm{Re}[T1a_{2}]|$', analysis_dir / 'kde_T1a2.png')
    save_phase_kde_only(phase_masks, total_abs_real, '$|\\mathrm{Re}[T1a_{1}]|+|\\mathrm{Re}[T1a_{2}]|$', analysis_dir / 'kde_T1a_total.png')
    save_phase_mean_barplot(summary_rows, 'T1a1_mean', 'Mean $|\\mathrm{Re}[T1a_1]|$', analysis_dir / 'mean_T1a1_by_phase.png', 'Mean of $|\\mathrm{Re}[T1a_1]|$ by phase')
    save_phase_mean_barplot(summary_rows, 'T1a2_mean', 'Mean $|\\mathrm{Re}[T1a_2]|$', analysis_dir / 'mean_T1a2_by_phase.png', 'Mean of $|\\mathrm{Re}[T1a_2]|$ by phase')
    save_phase_mean_barplot(summary_rows, 'Total_mean', 'Mean total intensity', analysis_dir / 'mean_total_by_phase.png', 'Mean of $|\\mathrm{Re}[T1a_1]| + |\\mathrm{Re}[T1a_2]|$ by phase')
    for phase_label, phase_mask in phase_masks.items():
        phase_tag = 'S_eq_0' if '0$' in phase_label and '-0.5' not in phase_label and ('-1/2' not in phase_label) else 'S_eq_mhalf'
        save_phase_masked_heatmap(alpha_vec, beta_vec, T1a1_abs_real, phase_mask, f'$|\\mathrm{{Re}}[T1a_1]|$ in {phase_label}', analysis_dir / f'T1a1_masked_{phase_tag}.png')
        save_phase_masked_heatmap(alpha_vec, beta_vec, T1a2_abs_real, phase_mask, f'$|\\mathrm{{Re}}[T1a_2]|$ in {phase_label}', analysis_dir / f'T1a2_masked_{phase_tag}.png')
        save_phase_masked_heatmap(alpha_vec, beta_vec, total_abs_real, phase_mask, f'$|\\mathrm{{Re}}[T1a_1]| + |\\mathrm{{Re}}[T1a_2]|$ in {phase_label}', analysis_dir / f'T1a_total_masked_{phase_tag}.png')
    summary_csv = tables_dir / 'phase_shg_summary.csv'
    distribution_csv = tables_dir / 'phase_distribution_stats.csv'
    summary_only_csv = tables_dir / 'phase_summary_statistics.csv'
    summary_only_xlsx = tables_dir / 'phase_summary_statistics.xlsx'
    phase_points_csv = tables_dir / 'phase_point_membership.csv'
    phase_points_xlsx = tables_dir / 'phase_point_membership.xlsx'
    write_records_csv(summary_rows, summary_csv)
    write_records_csv(summary_rows, summary_only_csv)
    write_records_csv(distribution_rows, distribution_csv)
    write_phase_summary_workbook(summary_rows, summary_only_xlsx)
    phase_point_rows = build_phase_point_rows(phase_masks, alpha_vec, beta_vec, average_shift_grid, T1a1_abs_real, T1a2_abs_real)
    write_records_csv(phase_point_rows, phase_points_csv)
    write_phase_point_workbook(phase_point_rows, phase_points_xlsx)
    t1a1_table = grid_to_table(alpha_vec, beta_vec, np.abs(np.real(shg_maps['T1a_1'])).T)
    t1a2_table = grid_to_table(alpha_vec, beta_vec, np.abs(np.real(shg_maps['T1a_2'])).T)
    t1a_total_table = grid_to_table(alpha_vec, beta_vec, (np.abs(np.real(shg_maps['T1a_1'])) + np.abs(np.real(shg_maps['T1a_2']))).T)
    avg_shift_table = grid_to_table(alpha_vec, beta_vec, average_shift_grid)
    rel_shift_table = grid_to_table(alpha_vec, beta_vec, relative_shift_grid)
    vort_table = grid_to_table(alpha_vec, beta_vec, minus_vortx_grid)
    rtpv_table = grid_to_table(alpha_vec, beta_vec, rtpv_grid)
    S0_table = grid_to_table(alpha_vec, beta_vec, S_0_grid)
    Spi_table = grid_to_table(alpha_vec, beta_vec, S_pi_grid)
    write_grid_csv(t1a1_table, tables_dir / 'T1a1_intensity_grid.csv')
    write_grid_csv(t1a2_table, tables_dir / 'T1a2_intensity_grid.csv')
    write_grid_csv(t1a_total_table, tables_dir / 'T1a_total_intensity_grid.csv')
    write_grid_csv(avg_shift_table, tables_dir / 'average_shift_grid.csv')
    write_grid_csv(rel_shift_table, tables_dir / 'relative_shift_grid.csv')
    write_grid_csv(vort_table, tables_dir / 'minus_vorticity_x_grid.csv')
    write_grid_csv(rtpv_table, tables_dir / 'rtp_v_grid.csv')
    write_grid_csv(S0_table, tables_dir / 'S_0_grid.csv')
    write_grid_csv(Spi_table, tables_dir / 'S_pi_grid.csv')
    if HAVE_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.title = 'README'
        ws['A1'] = 'topophase + full SHG split-terms summary'
        add_records_sheet(wb, 'phase_summary', summary_rows)
        add_records_sheet(wb, 'distribution_stats', distribution_rows)
        add_grid_sheet(wb, 'T1a1_grid', t1a1_table)
        add_grid_sheet(wb, 'T1a2_grid', t1a2_table)
        add_grid_sheet(wb, 'T1a_total_grid', t1a_total_table)
        add_grid_sheet(wb, 'avg_shift', avg_shift_table)
        add_grid_sheet(wb, 'relative_shift', rel_shift_table)
        add_grid_sheet(wb, 'minus_vortx', vort_table)
        add_grid_sheet(wb, 'rtp_v', rtpv_table)
        add_grid_sheet(wb, 'S_0', S0_table)
        add_grid_sheet(wb, 'S_pi', Spi_table)
        wb.save(outdir / 'phase_shg_summary.xlsx')
    if cfg.save_npy:
        np.save(outdir / 'alpha_vec.npy', alpha_vec)
        np.save(outdir / 'beta_vec.npy', beta_vec)
        np.save(outdir / 'invariants_grid.npy', invariants_grid)
        for k, v in shg_maps.items():
            np.save(outdir / f'{k}.npy', v)

def build_config() -> Config:
    return Config()

def main() -> None:
    cfg = build_config()
    topo_calc = TopoPhaseCalculator(cfg)
    shg_calc = FullSHGCalculator(cfg)
    alpha_vec = topo_calc.alpha_vec
    beta_vec = topo_calc.beta_vec
    full_tasks = [(ia, ib, float(a), float(b)) for ia, a in enumerate(alpha_vec) for ib, b in enumerate(beta_vec)]
    my_tasks = full_tasks[RANK::SIZE]
    barrier()
    local_topo: List[Tuple[int, int, np.ndarray]] = []
    local_shg: List[Tuple[int, int, Tuple[complex, ...]]] = []
    iterator = my_tasks
    for ia, ib, alpha, beta in iterator:
        topo_vals = topo_calc.calculate_invariants_for_point(alpha, beta)
        shg_vals = shg_calc.integrate_point(ia, ib, alpha, beta)
        local_topo.append((ia, ib, topo_vals))
        local_shg.append(shg_vals)
    all_topo = COMM.gather(local_topo, root=0) if COMM is not None else [local_topo]
    all_shg = COMM.gather(local_shg, root=0) if COMM is not None else [local_shg]
    if RANK != 0:
        return
    topo_results = topo_calc.gather_results(all_topo)
    shg_maps = shg_calc.gather_results(all_shg)
    T1a1_abs_real = np.abs(np.real(shg_maps['T1a_1']))
    T1a2_abs_real = np.abs(np.real(shg_maps['T1a_2']))
    total_abs_real = T1a1_abs_real + T1a2_abs_real
    phase_masks = make_phase_masks(topo_results['average_shift_grid'], cfg.phase_tol)
    summary_rows = summarize_by_phase(phase_masks, T1a1_abs_real, T1a2_abs_real)
    distribution_rows: List[Dict[str, float]] = []
    distribution_rows.extend(distribution_stats_by_phase(phase_masks, T1a1_abs_real, 'T1a_1_abs_real'))
    distribution_rows.extend(distribution_stats_by_phase(phase_masks, T1a2_abs_real, 'T1a_2_abs_real'))
    distribution_rows.extend(distribution_stats_by_phase(phase_masks, total_abs_real, 'T1a_total_abs_real'))
    save_outputs(cfg, topo_calc, shg_calc, topo_results, shg_maps, summary_rows, distribution_rows, phase_masks)
if __name__ == '__main__':
    main()
